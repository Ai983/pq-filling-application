from __future__ import annotations

from typing import Any, Dict, List, Optional, Tuple

from openpyxl.cell.cell import MergedCell

from app.core.config import (
    MAX_DOWN_SCAN,
    MAX_RIGHT_SCAN,
)
from app.core.constants import (
    CELL_TYPE_SECTION_HEADER,
    LAYOUT_BELOW_ANSWER,
    LAYOUT_BELOW_MERGED,
    LAYOUT_INLINE_MERGED_RIGHT,
    LAYOUT_INLINE_RIGHT,
    LAYOUT_ROW_TABLE,
    LAYOUT_UNKNOWN,
    RESOLVER_BELOW_CELL,
    RESOLVER_RIGHT_CELL,
    RESOLVER_RIGHT_MERGED,
    RESOLVER_ROW_TABLE,
    SKIP_REASON_AMBIGUOUS_TARGET,
    SKIP_REASON_EXISTING_VALUE,
    SKIP_REASON_FORMULA_CELL,
    SKIP_REASON_NO_TARGET,
    SKIP_REASON_UNSAFE_MERGED_TARGET,
)
from app.engine.utils import is_blank, normalize_text


# ---------------------------------------------------------------------
# Core merged-cell helpers
# ---------------------------------------------------------------------

def get_merged_anchor_cell(sheet, cell):
    """
    If the cell is part of a merged range, return the top-left anchor cell.
    Otherwise return the same cell.
    """
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
    return cell


def get_merged_range_string(sheet, cell) -> Optional[str]:
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return str(merged_range)
    return None


def is_formula_cell(cell) -> bool:
    value = getattr(cell, "value", None)
    return isinstance(value, str) and value.startswith("=")


def is_meaningful_value(value: Any) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    return text != ""


def is_writable_cell(sheet, cell) -> bool:
    """
    A writable target must resolve to a normal writable anchor cell.
    """
    anchor = get_merged_anchor_cell(sheet, cell)

    if isinstance(anchor, MergedCell):
        return False

    if is_formula_cell(anchor):
        return False

    return True


# ---------------------------------------------------------------------
# Cell role heuristics
# ---------------------------------------------------------------------

def _safe_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalized_text(value: Any) -> str:
    return normalize_text(_safe_text(value))


def _looks_like_label(text: str) -> bool:
    """
    Conservative label-likeness check used inside target resolution.
    """
    normalized = normalize_text(text)
    if not normalized:
        return False

    if len(normalized) > 140:
        return False

    label_markers = [
        "name",
        "company",
        "type",
        "address",
        "email",
        "mobile",
        "phone",
        "contact",
        "gst",
        "pan",
        "pf",
        "esi",
        "msme",
        "ifsc",
        "bank",
        "branch",
        "turnover",
        "project",
        "client",
        "location",
        "status",
        "value",
        "date",
        "designation",
        "details",
        "number",
        "registration",
        "certificate",
        "capacity",
        "staff",
        "manpower",
        "machinery",
        "office",
        "factory",
        "head office",
        "year of establishment",
        "type of company",
        "type of business",
    ]

    if normalized.endswith(":"):
        return True

    if any(marker == normalized or marker in normalized for marker in label_markers):
        return True

    return False


def _looks_like_structural_header(text: str) -> bool:
    normalized = normalize_text(text)
    if not normalized:
        return False

    header_markers = [
        "company credentials",
        "particulars",
        "details",
        "information",
        "annexure",
        "major projects executed",
        "projects under execution",
        "annual turnover",
        "capacity",
        "contact details",
        "list of certificates",
    ]
    return any(marker == normalized or marker in normalized for marker in header_markers)


def _has_visible_border(cell) -> bool:
    border = getattr(cell, "border", None)
    if not border:
        return False

    for side_name in ("left", "right", "top", "bottom"):
        side = getattr(border, side_name, None)
        if side and getattr(side, "style", None):
            return True
    return False


def _row_has_many_labels(sheet, row: int, max_scan: int = 8) -> bool:
    count = 0
    for col in range(1, min(sheet.max_column, max_scan) + 1):
        text = _safe_text(sheet.cell(row=row, column=col).value)
        if text and _looks_like_label(text):
            count += 1
    return count >= 3


def _nearby_right_has_another_label(sheet, label_row: int, label_col: int, target_col: int) -> bool:
    """
    Detects classic two-pair row pattern:
    [label][answer][label][answer]
    If there is another label between current label and target, penalize/skip.
    """
    for col in range(label_col + 1, target_col):
        text = _safe_text(sheet.cell(row=label_row, column=col).value)
        if text and _looks_like_label(text):
            return True
    return False


def _merged_span_width(sheet, cell) -> int:
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return int(merged_range.max_col - merged_range.min_col + 1)
    return 1


def _cell_layout_features(sheet, cell) -> Dict[str, Any]:
    anchor = get_merged_anchor_cell(sheet, cell)
    return {
        "coordinate": anchor.coordinate,
        "row": anchor.row,
        "column": anchor.column,
        "value": anchor.value,
        "text": _safe_text(anchor.value),
        "is_blank": is_blank(anchor.value),
        "is_formula": is_formula_cell(anchor),
        "is_merged": get_merged_range_string(sheet, anchor) is not None,
        "merged_range": get_merged_range_string(sheet, anchor),
        "merged_width": _merged_span_width(sheet, anchor),
        "has_border": _has_visible_border(anchor),
    }


# ---------------------------------------------------------------------
# Legacy candidate generation
# ---------------------------------------------------------------------

def get_candidate_target_cells(sheet, row: int, column: int):
    candidates = []

    # same row, right side
    for offset in range(1, MAX_RIGHT_SCAN + 1):
        if column + offset <= sheet.max_column:
            candidates.append(sheet.cell(row=row, column=column + offset))

    # below
    below_coords = [
        (row + 1, column),
        (row + 1, column + 1),
        (row + 1, column + 2),
        (row + 2, column),
        (row + 2, column + 1),
    ]
    for r, c in below_coords:
        if r <= sheet.max_row and c <= sheet.max_column:
            candidates.append(sheet.cell(row=r, column=c))

    return candidates


def score_target_cell(sheet, cell, label_row: int, label_col: int, label_text: Optional[str] = None) -> int:
    anchor = get_merged_anchor_cell(sheet, cell)

    if not is_writable_cell(sheet, cell):
        return -999

    if not is_blank(anchor.value):
        return -999

    text = _safe_text(anchor.value)
    if text and (_looks_like_label(text) or _looks_like_structural_header(text)):
        return -999

    row_diff = abs(anchor.row - label_row)
    col_diff = abs(anchor.column - label_col)

    score = 0

    if row_diff == 0 and col_diff == 1:
        score = 100
    elif row_diff == 0 and col_diff == 2:
        score = 92
    elif row_diff == 0 and col_diff == 3:
        score = 84
    elif row_diff == 0 and col_diff == 4:
        score = 74
    elif row_diff == 1 and col_diff == 0:
        score = 64
    elif row_diff == 1 and col_diff == 1:
        score = 60
    elif row_diff == 1 and col_diff == 2:
        score = 52
    elif row_diff == 2 and col_diff <= 1:
        score = 40
    else:
        score = 8

    if get_merged_range_string(sheet, anchor):
        score += 6

    if _has_visible_border(anchor):
        score += 2

    if row_diff == 0 and _nearby_right_has_another_label(sheet, label_row, label_col, anchor.column):
        score -= 18

    return score


def choose_best_target_cell(sheet, row: int, column: int):
    """
    Backward-compatible simple resolver.
    Returns (best_cell, score)
    """
    resolution = resolve_target_cell(
        sheet=sheet,
        label_row=row,
        label_col=column,
    )

    if not resolution["target_cell"]:
        return None, 0

    return resolution["target_cell"], int(resolution["score"])


def choose_table_value_cell(sheet, row: int, label_col: int):
    """
    Safer target resolver for horizontal row labels like:
    Engineers | 10
    FY 2023-24 | 12 Cr
    """
    resolution = resolve_table_value_cell(
        sheet=sheet,
        row=row,
        label_col=label_col,
    )

    if not resolution["target_cell"]:
        return None, 0

    return resolution["target_cell"], int(resolution["score"])


# ---------------------------------------------------------------------
# New structured resolver helpers
# ---------------------------------------------------------------------

def _build_result(
    sheet,
    label_row: int,
    label_col: int,
    anchor_cell=None,
    score: int = 0,
    layout_type: str = LAYOUT_UNKNOWN,
    resolver: str = "",
    reason: str = "",
) -> Dict[str, Any]:
    merged_range = None
    target_cell = None

    if anchor_cell is not None:
        target_cell = anchor_cell
        merged_range = get_merged_range_string(sheet, anchor_cell)

    return {
        "label_row": label_row,
        "label_col": label_col,
        "target_cell": target_cell,
        "target_coordinate": target_cell.coordinate if target_cell else None,
        "target_merged_range": merged_range,
        "score": int(score),
        "layout_confidence": round(max(0.0, min(score / 100.0, 1.0)), 4),
        "layout_type": layout_type,
        "resolver": resolver,
        "reason": reason,
    }


def _is_safe_target(sheet, cell) -> Tuple[bool, str]:
    anchor = get_merged_anchor_cell(sheet, cell)

    if isinstance(anchor, MergedCell):
        return False, SKIP_REASON_UNSAFE_MERGED_TARGET

    if is_formula_cell(anchor):
        return False, SKIP_REASON_FORMULA_CELL

    if not is_blank(anchor.value):
        return False, SKIP_REASON_EXISTING_VALUE

    return True, ""


def _unique_anchor_candidates(sheet, candidates: List[Any]) -> List[Any]:
    seen = set()
    unique = []

    for cell in candidates:
        anchor = get_merged_anchor_cell(sheet, cell)
        if anchor.coordinate in seen:
            continue
        seen.add(anchor.coordinate)
        unique.append(anchor)

    return unique


def _candidate_penalties(
    sheet,
    label_row: int,
    label_col: int,
    anchor,
    label_text: Optional[str] = None,
) -> int:
    penalty = 0

    anchor_text = _safe_text(anchor.value)
    if anchor_text and _looks_like_label(anchor_text):
        penalty += 30

    if anchor_text and _looks_like_structural_header(anchor_text):
        penalty += 30

    if anchor.row == label_row and _nearby_right_has_another_label(sheet, label_row, label_col, anchor.column):
        penalty += 16

    # avoid jumping too far in crowded rows
    if anchor.row == label_row and _row_has_many_labels(sheet, label_row) and (anchor.column - label_col) >= 4:
        penalty += 10

    # For compact identity rows, prefer immediate merged answer block
    normalized_label = normalize_text(label_text or "")
    if normalized_label and anchor.row == label_row:
        if any(
            key in normalized_label
            for key in [
                "company",
                "type",
                "address",
                "gst",
                "pan",
                "pf",
                "esi",
                "msme",
                "name",
                "registration",
                "year of establishment",
            ]
        ):
            if (anchor.column - label_col) > 3:
                penalty += 8

    return penalty


# ---------------------------------------------------------------------
# Resolver 1: same-row right-side cells
# ---------------------------------------------------------------------

def _resolve_inline_right(sheet, label_row: int, label_col: int, label_text: Optional[str] = None) -> Dict[str, Any]:
    candidates = []
    for offset in range(1, MAX_RIGHT_SCAN + 1):
        c = label_col + offset
        if c > sheet.max_column:
            break
        candidates.append(sheet.cell(row=label_row, column=c))

    anchors = _unique_anchor_candidates(sheet, candidates)
    scored: List[Tuple[int, Any, str, str]] = []

    for anchor in anchors:
        safe, reason = _is_safe_target(sheet, anchor)
        if not safe:
            continue

        col_diff = anchor.column - label_col
        text = _safe_text(anchor.value)
        if text and (_looks_like_label(text) or _looks_like_structural_header(text)):
            continue

        merged_range = get_merged_range_string(sheet, anchor)
        has_border = _has_visible_border(anchor)
        merged_width = _merged_span_width(sheet, anchor)

        if merged_range and col_diff == 1:
            score = 99
            layout_type = LAYOUT_INLINE_MERGED_RIGHT
            resolver = RESOLVER_RIGHT_MERGED
        elif merged_range and col_diff == 2:
            score = 95
            layout_type = LAYOUT_INLINE_MERGED_RIGHT
            resolver = RESOLVER_RIGHT_MERGED
        elif merged_range and col_diff <= 4:
            score = max(74, 96 - (col_diff * 5))
            layout_type = LAYOUT_INLINE_MERGED_RIGHT
            resolver = RESOLVER_RIGHT_MERGED
        else:
            if col_diff == 1:
                score = 94
            elif col_diff == 2:
                score = 88
            elif col_diff == 3:
                score = 80
            elif col_diff == 4:
                score = 68
            else:
                score = 56

            layout_type = LAYOUT_INLINE_RIGHT
            resolver = RESOLVER_RIGHT_CELL

        if merged_width >= 2:
            score += 4

        if has_border:
            score += 4

        score -= _candidate_penalties(
            sheet=sheet,
            label_row=label_row,
            label_col=label_col,
            anchor=anchor,
            label_text=label_text,
        )

        if score > 0:
            scored.append((score, anchor, layout_type, resolver))

    if not scored:
        return _build_result(
            sheet=sheet,
            label_row=label_row,
            label_col=label_col,
            anchor_cell=None,
            score=0,
            layout_type=LAYOUT_UNKNOWN,
            resolver=RESOLVER_RIGHT_CELL,
            reason=SKIP_REASON_NO_TARGET,
        )

    scored.sort(key=lambda item: item[0], reverse=True)
    best_score, best_anchor, layout_type, resolver = scored[0]

    return _build_result(
        sheet=sheet,
        label_row=label_row,
        label_col=label_col,
        anchor_cell=best_anchor,
        score=best_score,
        layout_type=layout_type,
        resolver=resolver,
        reason="",
    )


# ---------------------------------------------------------------------
# Resolver 2: below-cell / below-merged
# ---------------------------------------------------------------------

def _resolve_below(sheet, label_row: int, label_col: int, label_text: Optional[str] = None) -> Dict[str, Any]:
    candidates = []

    for row_offset in range(1, MAX_DOWN_SCAN + 1):
        for col_offset in range(0, 4):
            r = label_row + row_offset
            c = label_col + col_offset
            if r <= sheet.max_row and c <= sheet.max_column:
                candidates.append(sheet.cell(row=r, column=c))

    anchors = _unique_anchor_candidates(sheet, candidates)
    scored: List[Tuple[int, Any, str, str]] = []

    for anchor in anchors:
        safe, _ = _is_safe_target(sheet, anchor)
        if not safe:
            continue

        row_diff = anchor.row - label_row
        col_diff = abs(anchor.column - label_col)
        text = _safe_text(anchor.value)

        if text and (_looks_like_label(text) or _looks_like_structural_header(text)):
            continue

        merged_range = get_merged_range_string(sheet, anchor)
        has_border = _has_visible_border(anchor)

        if row_diff == 1 and col_diff == 0:
            score = 78
        elif row_diff == 1 and col_diff == 1:
            score = 74
        elif row_diff == 1 and col_diff == 2:
            score = 68
        elif row_diff == 1 and col_diff == 3:
            score = 62
        elif row_diff == 2 and col_diff == 0:
            score = 58
        elif row_diff == 2 and col_diff == 1:
            score = 54
        else:
            score = 38

        if merged_range:
            score += 8
            layout_type = LAYOUT_BELOW_MERGED
        else:
            layout_type = LAYOUT_BELOW_ANSWER

        if has_border:
            score += 4

        score -= _candidate_penalties(
            sheet=sheet,
            label_row=label_row,
            label_col=label_col,
            anchor=anchor,
            label_text=label_text,
        )

        if score > 0:
            scored.append((score, anchor, layout_type, RESOLVER_BELOW_CELL))

    if not scored:
        return _build_result(
            sheet=sheet,
            label_row=label_row,
            label_col=label_col,
            anchor_cell=None,
            score=0,
            layout_type=LAYOUT_UNKNOWN,
            resolver=RESOLVER_BELOW_CELL,
            reason=SKIP_REASON_NO_TARGET,
        )

    scored.sort(key=lambda item: item[0], reverse=True)
    best_score, best_anchor, layout_type, resolver = scored[0]

    return _build_result(
        sheet=sheet,
        label_row=label_row,
        label_col=label_col,
        anchor_cell=best_anchor,
        score=best_score,
        layout_type=layout_type,
        resolver=resolver,
        reason="",
    )


# ---------------------------------------------------------------------
# Resolver 3: row-table value resolver
# ---------------------------------------------------------------------

def resolve_table_value_cell(sheet, row: int, label_col: int) -> Dict[str, Any]:
    candidates = []

    for offset in range(1, min(MAX_RIGHT_SCAN, 6) + 1):
        c = label_col + offset
        if c <= sheet.max_column:
            candidates.append(sheet.cell(row=row, column=c))

    anchors = _unique_anchor_candidates(sheet, candidates)
    scored: List[Tuple[int, Any]] = []

    for idx, anchor in enumerate(anchors, start=1):
        safe, _ = _is_safe_target(sheet, anchor)
        if not safe:
            continue

        text = _safe_text(anchor.value)
        if text and (_looks_like_label(text) or _looks_like_structural_header(text)):
            continue

        col_diff = anchor.column - label_col

        if col_diff == 1:
            score = 100
        elif col_diff == 2:
            score = 94
        elif col_diff == 3:
            score = 86
        elif col_diff == 4:
            score = 76
        else:
            score = 64

        if get_merged_range_string(sheet, anchor):
            score += 4
        if _has_visible_border(anchor):
            score += 2

        if _nearby_right_has_another_label(sheet, row, label_col, anchor.column):
            score -= 20

        if score > 0:
            scored.append((score, anchor))

    if not scored:
        return _build_result(
            sheet=sheet,
            label_row=row,
            label_col=label_col,
            anchor_cell=None,
            score=0,
            layout_type=LAYOUT_ROW_TABLE,
            resolver=RESOLVER_ROW_TABLE,
            reason=SKIP_REASON_NO_TARGET,
        )

    scored.sort(key=lambda item: item[0], reverse=True)
    best_score, best_anchor = scored[0]

    return _build_result(
        sheet=sheet,
        label_row=row,
        label_col=label_col,
        anchor_cell=best_anchor,
        score=best_score,
        layout_type=LAYOUT_ROW_TABLE,
        resolver=RESOLVER_ROW_TABLE,
        reason="",
    )


# ---------------------------------------------------------------------
# Main ranked resolver
# ---------------------------------------------------------------------

def resolve_target_cell(
    sheet,
    label_row: int,
    label_col: int,
    label_text: Optional[str] = None,
    cell_type: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Ranked deterministic target-cell resolution.

    Order:
    1. right-side inline / merged block
    2. below-cell / below-merged
    3. fallback simple scoring over candidate neighborhood

    Returns structured metadata for logging and filler logic.
    """
    if cell_type == CELL_TYPE_SECTION_HEADER:
        return _build_result(
            sheet=sheet,
            label_row=label_row,
            label_col=label_col,
            anchor_cell=None,
            score=0,
            layout_type=LAYOUT_UNKNOWN,
            resolver="none",
            reason=SKIP_REASON_NO_TARGET,
        )

    results = [
        _resolve_inline_right(sheet, label_row, label_col, label_text=label_text),
        _resolve_below(sheet, label_row, label_col, label_text=label_text),
    ]

    # Legacy fallback scoring
    legacy_candidates = get_candidate_target_cells(sheet, label_row, label_col)
    legacy_scored = []
    seen = set()

    for cell in legacy_candidates:
        anchor = get_merged_anchor_cell(sheet, cell)
        if anchor.coordinate in seen:
            continue
        seen.add(anchor.coordinate)

        score = score_target_cell(
            sheet=sheet,
            cell=cell,
            label_row=label_row,
            label_col=label_col,
            label_text=label_text,
        )
        if score > 0:
            legacy_scored.append((score, anchor))

    if legacy_scored:
        legacy_scored.sort(key=lambda item: item[0], reverse=True)
        best_score, best_anchor = legacy_scored[0]
        results.append(
            _build_result(
                sheet=sheet,
                label_row=label_row,
                label_col=label_col,
                anchor_cell=best_anchor,
                score=best_score,
                layout_type=LAYOUT_INLINE_RIGHT if best_anchor.row == label_row else LAYOUT_BELOW_ANSWER,
                resolver="legacy_fallback",
                reason="",
            )
        )

    valid_results = [item for item in results if item.get("target_cell") is not None]
    if not valid_results:
        return _build_result(
            sheet=sheet,
            label_row=label_row,
            label_col=label_col,
            anchor_cell=None,
            score=0,
            layout_type=LAYOUT_UNKNOWN,
            resolver="none",
            reason=SKIP_REASON_NO_TARGET,
        )

    valid_results.sort(key=lambda item: item["score"], reverse=True)
    best = valid_results[0]

    # Ambiguity check:
    # if top two are very close but different cells, avoid risky fill
    if len(valid_results) > 1:
        second = valid_results[1]
        if (
            best["target_coordinate"] != second["target_coordinate"]
            and abs(best["score"] - second["score"]) <= 3
            and best["score"] < 95
        ):
            return _build_result(
                sheet=sheet,
                label_row=label_row,
                label_col=label_col,
                anchor_cell=None,
                score=0,
                layout_type=LAYOUT_UNKNOWN,
                resolver="ranked_conflict",
                reason=SKIP_REASON_AMBIGUOUS_TARGET,
            )

    return best