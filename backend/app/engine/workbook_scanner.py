from __future__ import annotations

from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from app.core.constants import (
    CELL_TYPE_EMPTY,
    CELL_TYPE_SECTION_HEADER,
    CELL_TYPE_UNKNOWN,
    DEFAULT_MAX_RIGHT_SCAN,
    SECTION_UNKNOWN,
)
from app.engine.field_classifier import classify_cell_text
from app.engine.section_resolver import resolve_section_from_text
from app.engine.utils import is_likely_label, normalize_text


def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _build_merged_lookup(sheet) -> Tuple[Dict[str, str], Dict[str, str]]:
    """
    Returns:
        merged_anchor_map:
            every cell coordinate in a merged range -> anchor coordinate
        merged_range_map:
            every cell coordinate in a merged range -> merged range string
    """
    merged_anchor_map: Dict[str, str] = {}
    merged_range_map: Dict[str, str] = {}

    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        anchor = sheet.cell(row=min_row, column=min_col).coordinate
        range_str = str(merged_range)

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                coord = sheet.cell(row=row, column=col).coordinate
                merged_anchor_map[coord] = anchor
                merged_range_map[coord] = range_str

    return merged_anchor_map, merged_range_map


def _has_visible_border(cell) -> bool:
    border = getattr(cell, "border", None)
    if not border:
        return False

    for side_name in ("left", "right", "top", "bottom"):
        side = getattr(border, side_name, None)
        if side and getattr(side, "style", None):
            return True
    return False


def _get_fill_hint(cell) -> Optional[str]:
    fill = getattr(cell, "fill", None)
    if not fill:
        return None

    fg = getattr(fill, "fgColor", None)
    if not fg:
        return None

    rgb = getattr(fg, "rgb", None)
    if rgb and rgb not in ("00000000", "000000", "FFFFFFFF", "FFFFFF"):
        return rgb

    indexed = getattr(fg, "indexed", None)
    if indexed is not None:
        return f"indexed:{indexed}"

    return None


def _cell_role_guess(
    text: str,
    normalized: str,
    cell_type: str,
    sheet,
    row_idx: int,
    col_idx: int,
) -> str:
    """
    Conservative role guess:
    - label
    - section_header
    - value
    - unknown
    """
    if not text:
        return CELL_TYPE_EMPTY

    if cell_type == CELL_TYPE_SECTION_HEADER:
        return CELL_TYPE_SECTION_HEADER

    if is_likely_label(text):
        return "label"

    # If right side is mostly blank, it may still behave like a label-ish prompt
    blank_right_count = 0
    for offset in range(1, DEFAULT_MAX_RIGHT_SCAN + 1):
        target_col = col_idx + offset
        if target_col > sheet.max_column:
            break
        right_value = _safe_str(sheet.cell(row=row_idx, column=target_col).value)
        if not right_value:
            blank_right_count += 1

    if blank_right_count >= 2 and len(normalized) <= 80:
        return "label"

    # Fall back to value
    return "value"


def _collect_neighbor_snapshot(sheet, row_idx: int, col_idx: int) -> Dict[str, Any]:
    def cell_text(r: int, c: int) -> str:
        if r < 1 or c < 1 or r > sheet.max_row or c > sheet.max_column:
            return ""
        return _safe_str(sheet.cell(row=r, column=c).value)

    return {
        "left": cell_text(row_idx, col_idx - 1),
        "right": cell_text(row_idx, col_idx + 1),
        "up": cell_text(row_idx - 1, col_idx),
        "down": cell_text(row_idx + 1, col_idx),
        "right_2": cell_text(row_idx, col_idx + 2),
        "down_2": cell_text(row_idx + 2, col_idx),
    }


def _is_sheet_meaningfully_empty(sheet) -> bool:
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                return False
    return True


def scan_workbook(file_path: str):
    """
    Scans workbook and returns:
        workbook,
        scan_result dict

    The scan_result is richer than the old implementation and is meant to support:
    - label matching
    - section resolution
    - merged-cell-safe writing
    - layout-aware target resolution
    - table detection

    Backward compatibility:
    - scan_result["findings"] keeps the earlier list-like shape so older code can still adapt.
    """
    wb = load_workbook(file_path)
    workbook_findings: List[Dict[str, Any]] = []
    sheets_data: List[Dict[str, Any]] = []

    for sheet in wb.worksheets:
        if _is_sheet_meaningfully_empty(sheet):
            sheets_data.append(
                {
                    "sheet_name": sheet.title,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "cells": [],
                    "labels": [],
                    "section_headers": [],
                    "merged_ranges": [],
                    "merged_anchor_map": {},
                    "merged_range_map": {},
                }
            )
            continue

        merged_anchor_map, merged_range_map = _build_merged_lookup(sheet)

        active_section = SECTION_UNKNOWN
        sheet_cells: List[Dict[str, Any]] = []
        sheet_labels: List[Dict[str, Any]] = []
        section_headers: List[Dict[str, Any]] = []

        for row in sheet.iter_rows():
            for cell in row:
                raw_text = _safe_str(cell.value)
                normalized = normalize_text(raw_text) if raw_text else ""

                is_empty = not raw_text
                merged_anchor = merged_anchor_map.get(cell.coordinate)
                merged_range = merged_range_map.get(cell.coordinate)

                # Basic styling / structure hints
                font = getattr(cell, "font", None)
                alignment = getattr(cell, "alignment", None)

                is_bold = bool(getattr(font, "bold", False))
                horizontal_alignment = getattr(alignment, "horizontal", None)
                vertical_alignment = getattr(alignment, "vertical", None)
                has_border = _has_visible_border(cell)
                fill_hint = _get_fill_hint(cell)
                has_formula = isinstance(cell.value, str) and str(cell.value).startswith("=")

                if is_empty:
                    cell_type = CELL_TYPE_EMPTY
                    section_for_cell = active_section
                    role_guess = CELL_TYPE_EMPTY
                else:
                    try:
                        cell_type = classify_cell_text(normalized)
                    except Exception:
                        cell_type = CELL_TYPE_UNKNOWN

                    role_guess = _cell_role_guess(
                        text=raw_text,
                        normalized=normalized,
                        cell_type=cell_type,
                        sheet=sheet,
                        row_idx=cell.row,
                        col_idx=cell.column,
                    )

                    # Section header detection
                    if cell_type == CELL_TYPE_SECTION_HEADER:
                        resolved_section = resolve_section_from_text(normalized)
                        if resolved_section != SECTION_UNKNOWN:
                            active_section = resolved_section
                        section_for_cell = active_section
                    else:
                        section_for_cell = active_section

                neighbor_snapshot = _collect_neighbor_snapshot(sheet, cell.row, cell.column)

                cell_data = {
                    "sheet": sheet.title,
                    "sheet_name": sheet.title,
                    "cell": cell.coordinate,
                    "coordinate": cell.coordinate,
                    "row": cell.row,
                    "column": cell.column,
                    "value": raw_text,
                    "raw_value": cell.value,
                    "normalized_value": normalized,
                    "cell_type": cell_type,
                    "role_guess": role_guess,
                    "active_section": section_for_cell,
                    "is_empty": is_empty,
                    "is_bold": is_bold,
                    "has_border": has_border,
                    "fill_hint": fill_hint,
                    "horizontal_alignment": horizontal_alignment,
                    "vertical_alignment": vertical_alignment,
                    "has_formula": has_formula,
                    "is_merged": merged_anchor is not None,
                    "merged_anchor": merged_anchor,
                    "merged_range": merged_range,
                    "neighbor_snapshot": neighbor_snapshot,
                }

                sheet_cells.append(cell_data)

                # Keep backward compatible "findings"
                if not is_empty and is_likely_label(raw_text):
                    workbook_findings.append(
                        {
                            "sheet": sheet.title,
                            "cell": cell.coordinate,
                            "value": raw_text,
                            "normalized_value": normalized,
                            "row": cell.row,
                            "column": cell.column,
                            "cell_type": cell_type,
                            "active_section": section_for_cell,
                            "is_bold": is_bold,
                            "has_border": has_border,
                            "is_merged": merged_anchor is not None,
                            "merged_anchor": merged_anchor,
                            "merged_range": merged_range,
                            "neighbor_snapshot": neighbor_snapshot,
                        }
                    )
                    sheet_labels.append(cell_data)

                if cell_type == CELL_TYPE_SECTION_HEADER:
                    section_headers.append(cell_data)

        sheets_data.append(
            {
                "sheet_name": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "cells": sheet_cells,
                "labels": sheet_labels,
                "section_headers": section_headers,
                "merged_ranges": [str(rng) for rng in sheet.merged_cells.ranges],
                "merged_anchor_map": merged_anchor_map,
                "merged_range_map": merged_range_map,
            }
        )

    scan_result = {
        "findings": workbook_findings,   # backward compatibility
        "sheets": sheets_data,
        "sheet_names": [sheet.title for sheet in wb.worksheets],
    }

    return wb, scan_result